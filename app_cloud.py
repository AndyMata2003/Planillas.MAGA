import streamlit as st
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder, DataReturnMode, GridUpdateMode
import base64
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import os
from weasyprint import HTML
import tempfile

# --- Función para generar PDF desde Excel en memoria ---
def generar_pdf_desde_excel(wb):
    ws = wb.active
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    df = pd.DataFrame(data)
    html = df.to_html(index=False)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmpfile:
        HTML(string=html).write_pdf(tmpfile.name)
        return tmpfile.name

# --- Utilidades para Streamlit Cloud ---
def load_font_base64(font_path):
    with open(font_path, "rb") as f:
        return base64.b64encode(f.read()).decode()

def set_cell_value_safe(ws, row, col_letter, value):
    cell = ws[f"{col_letter}{row}"]
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            if cell.coordinate == merged_range.start_cell.coordinate:
                cell.value = value
            return
    cell.value = value

def insertar_logo(ws, imagen_path, col='A', fila=1):
    try:
        img = XLImage(imagen_path)
        img.width = 133
        img.height = 143
        img.anchor = f"{col}{fila}"
        ws.add_image(img)
    except Exception:
        pass  # Continúa aunque no haya imagen

# --- Interfaz principal ---
st.set_page_config(page_title="Planillas MAGA Cloud", layout="wide")
st.title("🚀 Sistema de Planillas - Versión Cloud")

# Sidebar con configuración
with st.sidebar:
    st.header("Configuración de Oficio")
    unidad_ejecutora = st.text_input("Unidad Ejecutora", value="DAPCA")
    no_oficio = st.text_input("No. de Oficio", value="001")
    anio = st.text_input("Año", value="2025")
    codigo_oficio = f"{unidad_ejecutora.strip().upper()}-{no_oficio.strip().zfill(3)}-{anio.strip()}"
    st.markdown(f"**Código generado:** `{codigo_oficio}`")

# Uploaders fuera de condicionales anidados
uploaded_file_1 = st.file_uploader("Sube archivo con datos de comunidades", type=["xls", "xlsx"])
uploaded_file_2 = st.file_uploader("Sube archivo con beneficiarios", type=["xls", "xlsx"])

df_comunidades = pd.DataFrame()
df_beneficiarios = pd.DataFrame()

# Procesar archivo de comunidades
if uploaded_file_1:
    df_comunidades = pd.read_excel(uploaded_file_1)
    df_comunidades.columns = df_comunidades.columns.str.strip().str.replace('\xa0', '', regex=False)
    if 'Comunidad/ Establecimiento' in df_comunidades.columns:
        df_comunidades['Comunidad/ Establecimiento'] = df_comunidades['Comunidad/ Establecimiento'].astype(str).str.strip()
    else:
        st.error("El archivo de comunidades debe tener la columna 'Comunidad/ Establecimiento'")
        st.stop()

    st.success("Archivo de comunidades cargado")
    gb = GridOptionsBuilder.from_dataframe(df_comunidades)
    gb.configure_default_column(editable=True, filter=True, sortable=True)
    gridOptions = gb.build()
    grid_response = AgGrid(
        df_comunidades,
        gridOptions=gridOptions,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        update_mode=GridUpdateMode.MODEL_CHANGED,
        fit_columns_on_grid_load=True,
        height=300,
    )
    df_comunidades = pd.DataFrame(grid_response['data'])

# Procesar archivo de beneficiarios
if uploaded_file_2:
    df_beneficiarios = pd.read_excel(uploaded_file_2)
    df_beneficiarios.columns = df_beneficiarios.columns.str.strip()
    ref_col = next((col for col in df_beneficiarios.columns if col.strip().lower() == 'referencia'), None)
    if ref_col is None:
        st.error("No se encontró la columna 'Referencia' en el archivo de beneficiarios.")
        st.stop()
    df_beneficiarios[ref_col] = df_beneficiarios[ref_col].astype(str).str.strip().str.lower()

if not df_comunidades.empty and not df_beneficiarios.empty:
    comunidad_opciones = df_comunidades['Comunidad/ Establecimiento'].dropna().unique()
    comunidad_seleccionada = st.selectbox("Selecciona comunidad", options=comunidad_opciones)
    comunidad_seleccionada_lower = comunidad_seleccionada.strip().lower()
    df_filtrado = df_beneficiarios[df_beneficiarios[ref_col] == comunidad_seleccionada_lower]
    st.dataframe(df_filtrado)

    idx_comunidad = df_comunidades.index[
        df_comunidades['Comunidad/ Establecimiento'].str.strip().str.lower() == comunidad_seleccionada_lower
    ][0] + 1

    num_beneficiarios = len(df_filtrado)
    codigo_completo = f"{codigo_oficio}_CD{idx_comunidad}_P{num_beneficiarios}"
    st.sidebar.markdown(f"**Código completo:** `{codigo_completo}`")

    columnas_necesarias = ['PRIMER NOMBRE', 'SEGUNDO NOMBRE', 'TERCER NOMBRE', 'PRIMER APELLIDO', 'SEGUNDO APELLIDO', 'APELLIDO CASADA', 'CUI']
    df_temp = df_filtrado.copy()
    df_temp.columns = df_temp.columns.str.strip().str.upper()

    if all(col in df_temp.columns for col in columnas_necesarias):
        df_temp = df_temp[columnas_necesarias].fillna('').astype(str)
        df_temp['NOMBRE COMPLETO'] = (
            df_temp['PRIMER NOMBRE'] + ' ' + df_temp['SEGUNDO NOMBRE'] + ' ' + df_temp['TERCER NOMBRE'] + ' ' +
            df_temp['PRIMER APELLIDO'] + ' ' + df_temp['SEGUNDO APELLIDO'] + ' ' + df_temp['APELLIDO CASADA']
        ).str.replace(r'\s+', ' ', regex=True).str.strip()

        df_resultado = df_temp[['NOMBRE COMPLETO', 'CUI']]
        st.dataframe(df_resultado)

        if st.button("Generar Planilla Simplificada"):
            plantilla_path = "FormatoPlanillas.xlsx"
            if not os.path.exists(plantilla_path):
                st.error("El archivo 'FormatoPlanillas.xlsx' no se encuentra en el entorno.")
                st.stop()

            wb = load_workbook(plantilla_path)
            plantilla_hoja = wb["PLANILLAS"]
            bloques = [df_resultado.to_dict(orient='records')[i:i+10] for i in range(0, len(df_resultado), 10)]
            hojas_creadas = []

            for hoja_idx, bloque in enumerate(bloques):
                ws = wb.copy_worksheet(plantilla_hoja) if hoja_idx > 0 else plantilla_hoja
                ws.title = f"PLANILLA{hoja_idx+1}"
                insertar_logo(ws, "logo_maga.png")
                ws['C4'] = unidad_ejecutora
                ws['C7'] = df_comunidades.loc[idx_comunidad - 1, 'Departamento']
                ws['C9'] = df_comunidades.loc[idx_comunidad - 1, 'Municipio']
                ws['E9'] = comunidad_seleccionada
                ws['K1'] = codigo_completo

                for i, beneficiario in enumerate(bloque):
                    fila = 11 + i
                    set_cell_value_safe(ws, fila, 'B', beneficiario['NOMBRE COMPLETO'])
                    set_cell_value_safe(ws, fila, 'D', str(beneficiario['CUI']))

                hojas_creadas.append(ws.title)

            for hoja in wb.sheetnames:
                if hoja not in hojas_creadas:
                    del wb[hoja]

            pdf_path = generar_pdf_desde_excel(wb)
            with open(pdf_path, "rb") as f:
                st.download_button(
                    label="📄 Descargar Planilla en PDF",
                    data=f,
                    file_name=f"Planilla_{codigo_completo}.pdf",
                    mime="application/pdf"
                )
    else:
        st.warning("El archivo de beneficiarios no contiene todas las columnas necesarias para generar la planilla.")
