import streamlit as st
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder, DataReturnMode, GridUpdateMode
import base64
import openpyxl
import os
from win32com import client
import pythoncom
from openpyxl.drawing.image import Image as XLImage
import os

def obtener_ruta_planillas(dep, codigo_oficio):
    escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
    carpeta_base = os.path.join(escritorio, "PLANILLAS 2025")
    subcarpeta = f"Planillas_{dep.strip().replace(' ', '_')}_{codigo_oficio.strip()}"
    ruta_final = os.path.join(carpeta_base, subcarpeta)
    os.makedirs(ruta_final, exist_ok=True)
    return ruta_final

# Leer fuente TTF y convertir a base64
def load_font_base64(font_path):
    with open(font_path, "rb") as f:
        return base64.b64encode(f.read()).decode()

# Función segura para asignar valores en celdas (evita error en celdas combinadas)
def set_cell_value_safe(ws, row, col_letter, value):
    cell = ws[f"{col_letter}{row}"]
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            if cell.coordinate == merged_range.start_cell.coordinate:
                cell.value = value
            return
    cell.value = value

# Función para insertar el logo con dimensiones 3.51 cm × 3.77 cm (≈133×143 px)
def insertar_logo(ws, imagen_path, col='A', fila=1):
    img = XLImage(imagen_path)
    img.width = 133  # 3.51 cm
    img.height = 143  # 3.77 cm
    img.anchor = f"{col}{fila}"
    ws.add_image(img)

# Cargar fuente DAPCA
font_path = "fonts/DAPCA.ttf"
font_base64 = load_font_base64(font_path)

# Inyectar fuente en el estilo
st.markdown(f"""
    <style>
    @font-face {{
        font-family: 'DAPCA';
        src: url(data:font/ttf;base64,{font_base64}) format('truetype');
    }}
    .codigo-oficio {{
        font-family: 'DAPCA', sans-serif;
        font-size: 30px;
        font-weight: bold;
        color: #333;
        padding: 10px 0;
    }}
    </style>
""", unsafe_allow_html=True)

st.title("Sistema de Planillas")

# Sidebar: Unidad Ejecutora, No. de Oficio y Año
with st.sidebar:
    st.header("Configuración de Código")
    unidad_ejecutora = st.text_input("Unidad Ejecutora", value="DAPCA")
    no_oficio = st.text_input("No. de Oficio", value="001")
    anio = st.text_input("Año", value="2025")
    codigo_oficio = f"{unidad_ejecutora.strip().upper()}-{no_oficio.strip().zfill(3)}-{anio.strip()}"
    st.markdown(f'<div class="codigo-oficio">{codigo_oficio}</div>', unsafe_allow_html=True)

uploaded_file_1 = st.file_uploader("Sube archivo con datos de comunidades", type=["xls", "xlsx"])
df_comunidades = pd.DataFrame()

if uploaded_file_1:
    df_comunidades = pd.read_excel(uploaded_file_1)
    df_comunidades.columns = df_comunidades.columns.str.strip().str.replace('\xa0', '', regex=False)
    st.success("Archivo comunidades cargado")
    df_comunidades['Comunidad/ Establecimiento'] = df_comunidades['Comunidad/ Establecimiento'].astype(str).str.strip()

    gb = GridOptionsBuilder.from_dataframe(df_comunidades)
    gb.configure_default_column(editable=True, filter=True, sortable=True)
    gridOptions = gb.build()
    grid_response = AgGrid(
        df_comunidades,
        gridOptions=gridOptions,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        update_mode=GridUpdateMode.MODEL_CHANGED,
        fit_columns_on_grid_load=True,
        height=300,
    )
    df_comunidades = pd.DataFrame(grid_response['data'])

    uploaded_file_2 = st.file_uploader("Sube archivo con beneficiarios", type=["xls", "xlsx"])
    if uploaded_file_2:
        df_beneficiarios = pd.read_excel(uploaded_file_2)
        df_beneficiarios.columns = df_beneficiarios.columns.str.strip()
        st.success("Archivo beneficiarios cargado")

        ref_col = next((col for col in df_beneficiarios.columns if col.strip().lower() == 'referencia'), None)

        if ref_col is None:
            st.error("No se encontró la columna 'Referencia' en beneficiarios.")
        else:
            # Normalizar referencia en beneficiarios a minúsculas
            df_beneficiarios[ref_col] = df_beneficiarios[ref_col].astype(str).str.strip().str.lower()

            comunidad_opciones = df_comunidades['Comunidad/ Establecimiento'].dropna().unique()
            comunidad_seleccionada = st.selectbox("Selecciona comunidad para filtrar beneficiarios", options=comunidad_opciones)

            comunidad_seleccionada_lower = comunidad_seleccionada.strip().lower()
            df_filtrado = df_beneficiarios[df_beneficiarios[ref_col] == comunidad_seleccionada_lower]

            st.subheader(f"Beneficiarios en la comunidad: {comunidad_seleccionada}")
            st.dataframe(df_filtrado)

            try:
                idx_comunidad = df_comunidades.index[
                    df_comunidades['Comunidad/ Establecimiento'].str.strip().str.lower() == comunidad_seleccionada_lower
                ][0] + 1
            except IndexError:
                idx_comunidad = 1

            num_beneficiarios = len(df_filtrado)
            codigo_completo = f"{codigo_oficio}_CD{idx_comunidad}_P{num_beneficiarios}"

            st.sidebar.markdown("### Código Completo Generado")
            st.sidebar.markdown(f"**{codigo_completo}**")
            st.markdown("### Código completo generado:")
            st.markdown(f"**{codigo_completo}**")

            if not df_filtrado.empty:
                df_temp = df_filtrado.copy()
                df_temp.columns = df_temp.columns.str.strip().str.upper()
                columnas_necesarias = ['PRIMER NOMBRE', 'SEGUNDO NOMBRE', 'TERCER NOMBRE', 'PRIMER APELLIDO', 'SEGUNDO APELLIDO', 'APELLIDO CASADA', 'CUI']

                if all(col in df_temp.columns for col in columnas_necesarias):
                    df_nombres = df_temp[columnas_necesarias].copy()
                    for col in columnas_necesarias:
                        df_nombres[col] = df_nombres[col].fillna('').astype(str).str.strip()

                    df_nombres['NOMBRE COMPLETO'] = (
                        df_nombres['PRIMER NOMBRE'] + ' ' + df_nombres['SEGUNDO NOMBRE'] + ' ' +
                        df_nombres['TERCER NOMBRE'] + ' ' + df_nombres['PRIMER APELLIDO'] + ' ' +
                        df_nombres['SEGUNDO APELLIDO'] + ' ' + df_nombres['APELLIDO CASADA']
                    ).str.replace(r'\s+', ' ', regex=True).str.strip()

                    df_resultado = df_nombres[['NOMBRE COMPLETO', 'CUI']]

                    st.subheader("Listado de beneficiarios con nombre completo y CUI")
                    st.dataframe(df_resultado)
                    st.markdown(f"**Total de beneficiarios:** {len(df_resultado)}")
                    if st.button("Generar Planilla"):
                        try:
                            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
                            mun = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Municipio'].values[0]
                            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
                            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
                            Insu = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Insumo'].values[0]
                        except IndexError:
                            st.error("No se encontró la información completa para la comunidad seleccionada.")
                            st.stop()

                        plantilla = "FormatoPlanillas.xlsx"
                        logo_path = "logo_maga.png"

                        wb = openpyxl.load_workbook(plantilla)
                        plantilla_hoja = wb["PLANILLAS"]
                        beneficiarios = df_resultado.to_dict(orient='records')

                        bloques = [beneficiarios[i:i+10] for i in range(0, len(beneficiarios), 10)]
                        hojas_creadas = []

                        for hoja_idx, bloque in enumerate(bloques):
                            if hoja_idx == 0:
                                ws = plantilla_hoja
                                ws.title = f"PLANILLA{hoja_idx+1}"
                            else:
                                ws = wb.copy_worksheet(plantilla_hoja)
                                ws.title = f"PLANILLA{hoja_idx+1}"

                            insertar_logo(ws, logo_path)  # Aquí se usa con tamaño 133x143 px

                            # Mapeo de unidades ejecutoras a su nombre completo
                            unidad_map = {
                            "DAPA": "Departamento de: Apoyo a la Producción de Alimentos",
                            "DAU": "Departamento de: Agricultura Urbana",
                            "DADA": "Departamento de: Almacenamiento de Alimentos"
                            }
                            unidad_extra = unidad_map.get(unidad_ejecutora, unidad_ejecutora)
                            ws['C4'] = unidad_extra

                            ws['C7'] = dep
                            ws['C9'] = mun
                            ws['E9'] = comunidad_seleccionada
                            ws['K1'] = codigo_completo
                            ws['B23'] = tecnom
                            ws['B27'] = str(dpi)
                            ws['G24'] = Insu

                            for fila in range(1, 3435):
                                ws.row_dimensions[fila].hidden = True

                            fila_inicio = 1 + hoja_idx * 34
                            fila_fin = fila_inicio + 33
                            for fila in range(fila_inicio, fila_fin + 1):
                                ws.row_dimensions[fila].hidden = False

                            for i in range(10):
                                fila = fila_inicio + 11 + i
                                if i < len(bloque):
                                    set_cell_value_safe(ws, fila, 'B', bloque[i]['NOMBRE COMPLETO'])
                                    set_cell_value_safe(ws, fila, 'D', str(bloque[i]['CUI']))
                                    ws.row_dimensions[fila].hidden = False
                                else:
                                    set_cell_value_safe(ws, fila, 'B', "")
                                    set_cell_value_safe(ws, fila, 'D', "")
                                    ws.row_dimensions[fila].hidden = True

                            hojas_creadas.append(ws.title)

                        for hoja in wb.sheetnames:
                            if hoja not in hojas_creadas:
                                wb.remove(wb[hoja])

                        
                        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
                        nombre_archivo = f"{idx_comunidad} - {dep}, {mun}, PLANILLA.xlsx"
                        output_excel = os.path.join(ruta_guardado, nombre_archivo)
                        pdf_output = output_excel.replace(".xlsx", ".pdf")
                        wb.save(output_excel)

                        pythoncom.CoInitialize()
                        try:
                            excel = client.Dispatch("Excel.Application")
                            excel.Application.Visible = False
                            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
                            pdf_output = output_excel.replace(".xlsx", ".pdf")
                            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
                            wb_pdf.Close(False)
                            excel.Quit()
                        finally:
                            pythoncom.CoUninitialize()

                        with open(pdf_output, "rb") as f:
                            st.download_button(
                                label="Descargar Planilla PDF",
                                data=f,
                                file_name=os.path.basename(pdf_output),
                                mime="application/pdf"
                            )# Mostrar los botones adicionales solo si se ha cargado comunidades y seleccionado una comunidad
if uploaded_file_1 and 'comunidad_seleccionada' in locals():
    # Botón para Planilla Adicional (tipo entrega, sin beneficiarios)
    if st.button("Generar Planilla Adicional"):
        try:
            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
            Insu = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Insumo'].values[0]
        except IndexError:
            st.error("No se encontró la información completa para la comunidad seleccionada.")
            st.stop()

        plantilla = "FormatoPlanillas.xlsx"
        logo_path = "logo_maga.png"
        wb = openpyxl.load_workbook(plantilla)

        if "ADICIONALPLA" not in wb.sheetnames:
            st.error("La hoja 'ADICIONALPLA' no se encuentra en el archivo Excel.")
            st.stop()

        ws = wb["ADICIONALPLA"]
        ws.title = "ADICIONALPLA1"

        insertar_logo(ws, logo_path)

        unidad_map = {
        "DAPA": "Departamento de: Apoyo a la Producción de Alimentos",
        "DAU": "Departamento de: Agricultura Urbana",
        "DADA": "Departamento de: Almacenamiento de Alimentos"
        }
        unidad_extra = unidad_map.get(unidad_ejecutora, unidad_ejecutora)
        ws['C4'] = unidad_extra

        ws['C7'] = dep
        ws['K1'] = codigo_oficio
        ws['B23'] = tecnom
        ws['B27'] = str(dpi)
        ws['G24'] = Insu

        for hoja in wb.sheetnames:
            if hoja != "ADICIONALPLA1":
                wb.remove(wb[hoja])

        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
        nombre_archivo = f"{idx_comunidad} - {dep}, PLANILLA ADICIONAL.xlsx"
        output_excel = os.path.join(ruta_guardado, nombre_archivo)
        pdf_output = output_excel.replace(".xlsx", ".pdf")
        wb.save(output_excel)

        pythoncom.CoInitialize()
        try:
            excel = client.Dispatch("Excel.Application")
            excel.Application.Visible = False
            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
            wb_pdf.Sheets(1).Select()
            pdf_output = output_excel.replace(".xlsx", ".pdf")
            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
            wb_pdf.Close(False)
            excel.Quit()
        finally:
            pythoncom.CoUninitialize()

        with open(pdf_output, "rb") as f:
            st.download_button(
                label="Descargar Planilla Adicional PDF",
                data=f,
                file_name=os.path.basename(pdf_output),
                mime="application/pdf"
            )
if not df_comunidades.empty and uploaded_file_2:
    if st.button("Generar Planilla de Asistencia"):
        try:
            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
            mun = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Municipio'].values[0]
            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
        except IndexError:
            st.error("No se encontró la información completa para la comunidad seleccionada.")
            st.stop()

        plantilla = "FormatoPlanillas.xlsx"
        logo_path = "logo_maga.png"

        wb = openpyxl.load_workbook(plantilla)
        plantilla_hoja = wb["PLANILLASASIST"]
        beneficiarios = df_resultado.to_dict(orient='records')

        bloques = [beneficiarios[i:i+10] for i in range(0, len(beneficiarios), 10)]
        hojas_creadas = []

        for hoja_idx, bloque in enumerate(bloques):
            if hoja_idx == 0:
                ws = plantilla_hoja
                ws.title = f"PLANILLASASIST{hoja_idx+1}"
            else:
                ws = wb.copy_worksheet(plantilla_hoja)
                ws.title = f"PLANILLASASIST{hoja_idx+1}"

            insertar_logo(ws, logo_path)

            unidad_map = {
            "DAPA": "Departamento de: Apoyo a la Producción de Alimentos",
            "DAU": "Departamento de: Agricultura Urbana",
            "DADA": "Departamento de: Almacenamiento de Alimentos"
            }
            unidad_extra = unidad_map.get(unidad_ejecutora, unidad_ejecutora)
            ws['C4'] = unidad_extra

            ws['C7'] = dep
            ws['C9'] = mun
            ws['E9'] = comunidad_seleccionada
            ws['K1'] = codigo_completo
            ws['B23'] = tecnom
            ws['B27'] = str(dpi)
            ws['G24'] = "Asistencia Técnica"  # CAMBIO AQUÍ

            for fila in range(1, 1429):
                ws.row_dimensions[fila].hidden = True

            fila_inicio = 1 + hoja_idx * 34
            fila_fin = fila_inicio + 33
            for fila in range(fila_inicio, fila_fin + 1):
                ws.row_dimensions[fila].hidden = False

            for i in range(10):
                fila = fila_inicio + 11 + i
                if i < len(bloque):
                    set_cell_value_safe(ws, fila, 'B', bloque[i]['NOMBRE COMPLETO'])
                    set_cell_value_safe(ws, fila, 'D', str(bloque[i]['CUI']))
                    ws.row_dimensions[fila].hidden = False
                else:
                    set_cell_value_safe(ws, fila, 'B', "")
                    set_cell_value_safe(ws, fila, 'D', "")
                    ws.row_dimensions[fila].hidden = True

            hojas_creadas.append(ws.title)

        for hoja in wb.sheetnames:
            if hoja not in hojas_creadas:
                wb.remove(wb[hoja])

        
        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
        nombre_archivo = f"{idx_comunidad} - {dep}, {mun}, PLANILLA DE ASISTENCIA.xlsx"
        output_excel = os.path.join(ruta_guardado, nombre_archivo)
        pdf_output = output_excel.replace(".xlsx", ".pdf")
        wb.save(output_excel)

        pythoncom.CoInitialize()
        try:
            excel = client.Dispatch("Excel.Application")
            excel.Application.Visible = False
            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
            pdf_output = output_excel.replace(".xlsx", ".pdf")
            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
            wb_pdf.Close(False)
            excel.Quit()
        finally:
            pythoncom.CoUninitialize()

        with open(pdf_output, "rb") as f:
            st.download_button(
                label="Descargar Planilla de Asistencia PDF",
                data=f,
                file_name=os.path.basename(pdf_output),
                mime="application/pdf"
            )


if uploaded_file_1 and 'comunidad_seleccionada' in locals():
    if st.button("Generar Planilla de Asistencia Adicional"):
        try:
            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
        except IndexError:
            st.error("No se encontró la información completa para la comunidad seleccionada.")
            st.stop()

        plantilla = "FormatoPlanillas.xlsx"
        logo_path = "logo_maga.png"
        wb = openpyxl.load_workbook(plantilla)

        if "ASISTENCIAAD" not in wb.sheetnames:
            st.error("La hoja 'ASISTENCIAAD' no se encuentra en el archivo Excel.")
            st.stop()

        ws = wb["ASISTENCIAAD"]
        ws.title = "ASISTENCIAAD1"

        insertar_logo(ws, logo_path)  # Tamaño 133x143 px

        unidad_map = {
        "DAPA": "Departamento de: Apoyo a la Producción de Alimentos",
        "DAU": "Departamento de: Agricultura Urbana",
        "DADA": "Departamento de: Almacenamiento de Alimentos"
        }
        unidad_extra = unidad_map.get(unidad_ejecutora, unidad_ejecutora)
        ws['C4'] = unidad_extra

        ws['C7'] = dep
        ws['K1'] = codigo_oficio
        ws['B23'] = tecnom
        ws['B27'] = str(dpi)
        ws['G24'] = "Asistencia Técnica"

        for hoja in wb.sheetnames:
            if hoja != "ASISTENCIAAD1":
                wb.remove(wb[hoja])

        
        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
        nombre_archivo = f"{dep}, ASISTENCIA ADICIONAL.xlsx"
        output_excel = os.path.join(ruta_guardado, nombre_archivo)
        pdf_output = output_excel.replace(".xlsx", ".pdf")
        wb.save(output_excel)

        pythoncom.CoInitialize()
        try:
            excel = client.Dispatch("Excel.Application")
            excel.Application.Visible = False
            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
            wb_pdf.Sheets(1).Select()
            pdf_output = output_excel.replace(".xlsx", ".pdf")
            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
            wb_pdf.Close(False)
            excel.Quit()
        finally:
            pythoncom.CoUninitialize()

        with open(pdf_output, "rb") as f:
            st.download_button(
                label="Descargar Planilla de Asistencia Adicional PDF",
                data=f,
                file_name=os.path.basename(pdf_output),
                mime="application/pdf"
            )
if uploaded_file_1 and 'comunidad_seleccionada' in locals():
 if uploaded_file_1 and 'comunidad_seleccionada' in locals():
    if st.button("Generar Planilla de Capacitación"):
        try:
            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
            mun = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Municipio'].values[0]
            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
            Insu = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Insumo'].values[0]
            capa = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Listado de Registro de capacitacion y asistencia Tecnica'].values[0]
        except IndexError:
            st.error("No se encontró la información completa para la comunidad seleccionada.")
            st.stop()

        plantilla = "FormatoPlanillas.xlsx"
        logo_path = "logo_maga.png"
        wb = openpyxl.load_workbook(plantilla)
        plantilla_hoja = wb["LISTADOCAPA"]
        beneficiarios = df_resultado.to_dict(orient='records')

        bloques = [beneficiarios[i:i+10] for i in range(0, len(beneficiarios), 10)]
        hojas_creadas = []

        for hoja_idx, bloque in enumerate(bloques):
            ws = wb.copy_worksheet(plantilla_hoja)
            ws.title = f"PLANILLA{hoja_idx+1}"

            # Insertar logo en A2 con tamaño 2.08 cm x 2.00 cm (≈ 79 x 76 px)
            img = XLImage(logo_path)
            img.width = 79
            img.height = 76
            img.anchor = "A2"
            ws.add_image(img)

            # Mapeo de unidades ejecutoras a su nombre completo
            unidad_map = {
                "DAPA": "Departamento de: Apoyo a la Producción de Alimentos",
                "DAU": "Departamento de: Agricultura Urbana",
                "DADA": "Departamento de: Almacenamiento de Alimentos"
            }
            unidad_extra = unidad_map.get(unidad_ejecutora, unidad_ejecutora)
            ws['A5'] = unidad_extra

            # Encabezados
            ws['C11'] = dep
            ws['C12'] = mun
            ws['C13'] = comunidad_seleccionada
            ws['A4'] = codigo_completo
            ws['A5'] = unidad_extra
            ws['A32'] = tecnom
            ws['C14'] = capa

            # Ocultar todas las filas
            for fila in range(1, 2091):
                ws.row_dimensions[fila].hidden = True

            fila_inicio = 1 + hoja_idx * 41
            fila_fin = fila_inicio + 40
            for fila in range(fila_inicio, fila_fin + 1):
                ws.row_dimensions[fila].hidden = False

            for i in range(10):
                fila = fila_inicio + 17 + i
                if i < len(bloque):
                    set_cell_value_safe(ws, fila, 'B', bloque[i]['NOMBRE COMPLETO'])
                    set_cell_value_safe(ws, fila, 'D', str(bloque[i]['CUI']))
                    ws.row_dimensions[fila].hidden = False
                else:
                    set_cell_value_safe(ws, fila, 'B', "")
                    set_cell_value_safe(ws, fila, 'D', "")
                    ws.row_dimensions[fila].hidden = True

            hojas_creadas.append(ws.title)

        # Eliminar la hoja original de plantilla
        if "LISTADOCAPA" in wb.sheetnames:
            wb.remove(wb["LISTADOCAPA"])

        for hoja in wb.sheetnames:
            if hoja not in hojas_creadas:
                wb.remove(wb[hoja])

        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
        nombre_archivo = f"{idx_comunidad}-{dep}, {mun}, LISTADO DE CAPACITACIÓN.xlsx"
        output_excel = os.path.join(ruta_guardado, nombre_archivo)
        pdf_output = output_excel.replace(".xlsx", ".pdf")
        wb.save(output_excel)

        pythoncom.CoInitialize()
        try:
            excel = client.Dispatch("Excel.Application")
            excel.Application.Visible = False
            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
            wb_pdf.Sheets(1).Select()
            pdf_output = output_excel.replace(".xlsx", ".pdf")
            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
            wb_pdf.Close(False)
            excel.Quit()
        finally:
            pythoncom.CoUninitialize()

        with open(pdf_output, "rb") as f:
            st.download_button(
                label="Descargar Planilla de Capacitacion PDF",
                data=f,
                file_name=os.path.basename(pdf_output),
                mime="application/pdf"
            )

if uploaded_file_1 and 'comunidad_seleccionada' in locals():
    if st.button("Generar Planilla de Capacitacion Adicional"):
        try:
            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
            mun = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Municipio'].values[0]
            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
            Insu = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Insumo'].values[0]
        except IndexError:
            st.error("No se encontró la información completa para la comunidad seleccionada.")
            st.stop()

        plantilla = "FormatoPlanillas.xlsx"
        logo_path = "logo_maga.png"
        wb = openpyxl.load_workbook(plantilla)

        if "LISTADOCAPAAD" not in wb.sheetnames:
            st.error("La hoja 'LISTADOCAPAAD' no se encuentra en el archivo Excel.")
            st.stop()

        ws = wb["LISTADOCAPAAD"]
        ws.title = "LISTADOCAPAAD1"

        # Insertar logo
        img = XLImage(logo_path)
        img.width = 79
        img.height = 76
        img.anchor = "A2"
        ws.add_image(img)

        unidad_map = {
        "DAPA": "Departamento de: Apoyo a la Producción de Alimentos",
        "DAU": "Departamento de: Agricultura Urbana",
        "DADA": "Departamento de: Almacenamiento de Alimentos"
        }
        unidad_extra = unidad_map.get(unidad_ejecutora, unidad_ejecutora)
        ws['A5'] = unidad_extra

        ws['C11'] = dep
        ws['A4'] = codigo_oficio
        ws['A5'] = unidad_extra
        ws['A32'] = tecnom

        for hoja in wb.sheetnames:
            if hoja != "LISTADOCAPAAD1":
                wb.remove(wb[hoja])

        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
        nombre_archivo = f"{dep}, CAPACITACIÓN ADICIONAL.xlsx"
        output_excel = os.path.join(ruta_guardado, nombre_archivo)
        pdf_output = output_excel.replace(".xlsx", ".pdf")
        wb.save(output_excel)

        pythoncom.CoInitialize()
        try:
            excel = client.Dispatch("Excel.Application")
            excel.Application.Visible = False
            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
            wb_pdf.Sheets(1).Select()
            pdf_output = output_excel.replace(".xlsx", ".pdf")
            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
            wb_pdf.Close(False)
            excel.Quit()
        finally:
            pythoncom.CoUninitialize()

        with open(pdf_output, "rb") as f:
            st.download_button(
                label="Descargar Planilla de Asistencia Adicional PDF",
                data=f,
                file_name=os.path.basename(pdf_output),
                mime="application/pdf"
            )
if uploaded_file_1 and 'comunidad_seleccionada' in locals():
    if st.button("Generar Planilla de DAU"):
        try:
            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
            mun = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Municipio'].values[0]
            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
            Insu = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Insumo'].values[0]
            CoEs = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CODIGO ESCOLAR'].values[0]
        except IndexError:
            st.error("No se encontró la información completa para la comunidad seleccionada.")
            st.stop()

        plantilla = "FormatoPlanillas.xlsx"
        logo_path = "logo_maga.png"

        wb = openpyxl.load_workbook(plantilla)
        plantilla_hoja = wb["PLANILLASDAU"]
        beneficiarios = df_resultado.to_dict(orient='records')

        bloques = [beneficiarios[i:i+10] for i in range(0, len(beneficiarios), 10)]
        hojas_creadas = []

        for hoja_idx, bloque in enumerate(bloques):
            if hoja_idx == 0:
                ws = plantilla_hoja
                ws.title = f"PLANILLASDAU{hoja_idx+1}"
            else:
                ws = wb.copy_worksheet(plantilla_hoja)
                ws.title = f"PLANILLASDAU{hoja_idx+1}"

            insertar_logo(ws, logo_path)

        ws['C7'] = dep
        ws['C9'] = mun
        ws['E7'] = comunidad_seleccionada
        ws['K1'] = codigo_completo
        ws['B23'] = tecnom
        ws['B27'] = str(dpi)
        ws['J24'] = Insu
        ws['E9'] = CoEs

        for fila in range(1, 1701):
                ws.row_dimensions[fila].hidden = True

        fila_inicio = 1 + hoja_idx * 34
        fila_fin = fila_inicio + 33
        for fila in range(fila_inicio, fila_fin + 1):
                ws.row_dimensions[fila].hidden = False

        for i in range(10):
                fila = fila_inicio + 11 + i
                if i < len(bloque):
                    set_cell_value_safe(ws, fila, 'B', bloque[i]['NOMBRE COMPLETO'])
                    set_cell_value_safe(ws, fila, 'D', str(bloque[i]['CUI']))
                    ws.row_dimensions[fila].hidden = False
                else:
                    set_cell_value_safe(ws, fila, 'B', "")
                    set_cell_value_safe(ws, fila, 'D', "")
                    ws.row_dimensions[fila].hidden = True

        hojas_creadas.append(ws.title)

        for hoja in wb.sheetnames:
            if hoja not in hojas_creadas:
                wb.remove(wb[hoja])

        
        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
        nombre_archivo = f"{idx_comunidad} - {dep}, {mun}, PLANILLA DAU.xlsx"
        output_excel = os.path.join(ruta_guardado, nombre_archivo)
        pdf_output = output_excel.replace(".xlsx", ".pdf")
        wb.save(output_excel)

        pythoncom.CoInitialize()
        try:
            excel = client.Dispatch("Excel.Application")
            excel.Application.Visible = False
            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
            pdf_output = output_excel.replace(".xlsx", ".pdf")
            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
            wb_pdf.Close(False)
            excel.Quit()
        finally:
            pythoncom.CoUninitialize()

        with open(pdf_output, "rb") as f:
            st.download_button(
                label="Descargar Planilla de Asistencia PDF",
                data=f,
                file_name=os.path.basename(pdf_output),
                mime="application/pdf"
            )
if uploaded_file_1 and 'comunidad_seleccionada' in locals():
    if st.button("Generar Planilla Adicional DAU"):
        try:
            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
        except IndexError:
            st.error("No se encontró la información completa para la comunidad seleccionada.")
            st.stop()

        plantilla = "FormatoPlanillas.xlsx"
        logo_path = "logo_maga.png"
        wb = openpyxl.load_workbook(plantilla)

        if "PLANILLASDAUAD" not in wb.sheetnames:
            st.error("La hoja 'PLANILLASDAUAD' no se encuentra en el archivo Excel.")
            st.stop()

        ws = wb["PLANILLASDAUAD"]
        ws.title = "PLANILLASDAUAD1"

        insertar_logo(ws, logo_path)  # Tamaño 133x143 px

        ws['C7'] = dep
        ws['K1'] = codigo_oficio
        ws['B23'] = tecnom
        ws['B27'] = str(dpi)

        for hoja in wb.sheetnames:
            if hoja != "PLANILLASDAUAD1":
                wb.remove(wb[hoja])

        
        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
        nombre_archivo = f"{dep}, PLANILLA ADICIONAL DAU.xlsx"
        output_excel = os.path.join(ruta_guardado, nombre_archivo)
        pdf_output = output_excel.replace(".xlsx", ".pdf")
        wb.save(output_excel)

        pythoncom.CoInitialize()
        try:
            excel = client.Dispatch("Excel.Application")
            excel.Application.Visible = False
            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
            wb_pdf.Sheets(1).Select()
            pdf_output = output_excel.replace(".xlsx", ".pdf")
            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
            wb_pdf.Close(False)
            excel.Quit()
        finally:
            pythoncom.CoUninitialize()

        with open(pdf_output, "rb") as f:
            st.download_button(
                label="Descargar Planilla Adicional DAU PDF",
                data=f,
                file_name=os.path.basename(pdf_output),
                mime="application/pdf"
            )
if not df_comunidades.empty and uploaded_file_2:
    if st.button("Generar Planilla de Asistencia DAU"):
        try:
            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
            mun = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Municipio'].values[0]
            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
            CoEs = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CODIGO ESCOLAR'].values[0]
            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
        except IndexError:
            st.error("No se encontró la información completa para la comunidad seleccionada.")
            st.stop()

        plantilla = "FormatoPlanillas.xlsx"
        logo_path = "logo_maga.png"

        wb = openpyxl.load_workbook(plantilla)
        plantilla_hoja = wb["ASISTENCIADAU"]
        beneficiarios = df_resultado.to_dict(orient='records')

        bloques = [beneficiarios[i:i+10] for i in range(0, len(beneficiarios), 10)]
        hojas_creadas = []

        for hoja_idx, bloque in enumerate(bloques):
            if hoja_idx == 0:
                ws = plantilla_hoja
                ws.title = f"ASISTENCIADAU{hoja_idx+1}"
            else:
                ws = wb.copy_worksheet(plantilla_hoja)
                ws.title = f"ASISTENCIADAU{hoja_idx+1}"

            insertar_logo(ws, logo_path)

            ws['C7'] = dep
            ws['C9'] = mun
            ws['E7'] = comunidad_seleccionada
            ws['K1'] = codigo_completo
            ws['B23'] = tecnom
            ws['B27'] = str(dpi)
            ws['E9'] = CoEs
            ws['J24'] = "Asistencia Técnica"  # CAMBIO AQUÍ

            for fila in range(1, 1701):
                ws.row_dimensions[fila].hidden = True

            fila_inicio = 1 + hoja_idx * 34
            fila_fin = fila_inicio + 33
            for fila in range(fila_inicio, fila_fin + 1):
                ws.row_dimensions[fila].hidden = False

            for i in range(10):
                fila = fila_inicio + 11 + i
                if i < len(bloque):
                    set_cell_value_safe(ws, fila, 'B', bloque[i]['NOMBRE COMPLETO'])
                    set_cell_value_safe(ws, fila, 'D', str(bloque[i]['CUI']))
                    ws.row_dimensions[fila].hidden = False
                else:
                    set_cell_value_safe(ws, fila, 'B', "")
                    set_cell_value_safe(ws, fila, 'D', "")
                    ws.row_dimensions[fila].hidden = True

            hojas_creadas.append(ws.title)

        for hoja in wb.sheetnames:
            if hoja not in hojas_creadas:
                wb.remove(wb[hoja])

        
        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
        nombre_archivo = f"{idx_comunidad} - {dep}, {mun}, PLANILLA DE ASISTENCIA DAU.xlsx"
        output_excel = os.path.join(ruta_guardado, nombre_archivo)
        pdf_output = output_excel.replace(".xlsx", ".pdf")
        wb.save(output_excel)

        pythoncom.CoInitialize()
        try:
            excel = client.Dispatch("Excel.Application")
            excel.Application.Visible = False
            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
            pdf_output = output_excel.replace(".xlsx", ".pdf")
            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
            wb_pdf.Close(False)
            excel.Quit()
        finally:
            pythoncom.CoUninitialize()

        with open(pdf_output, "rb") as f:
            st.download_button(
                label="Descargar Planilla de Asistencia DAU PDF",
                data=f,
                file_name=os.path.basename(pdf_output),
                mime="application/pdf"
            )


if uploaded_file_1 and 'comunidad_seleccionada' in locals():
    if st.button("Generar Planilla de Asistencia Adicional DAU"):
        try:
            dep = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Departamento'].values[0]
            tecnom = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'Nombre del técnico'].values[0]
            dpi = df_comunidades.loc[df_comunidades['Comunidad/ Establecimiento'] == comunidad_seleccionada, 'CUI del técnico'].values[0]
        except IndexError:
            st.error("No se encontró la información completa para la comunidad seleccionada.")
            st.stop()

        plantilla = "FormatoPlanillas.xlsx"
        logo_path = "logo_maga.png"
        wb = openpyxl.load_workbook(plantilla)

        if "ASISTENCIADAUAD" not in wb.sheetnames:
            st.error("La hoja 'ASISTENCIADAUAD' no se encuentra en el archivo Excel.")
            st.stop()

        ws = wb["ASISTENCIADAUAD"]
        ws.title = "ASISTENCIADAUAD1"

        insertar_logo(ws, logo_path)  # Tamaño 133x143 px

        ws['C7'] = dep
        ws['K1'] = codigo_oficio
        ws['B23'] = tecnom
        ws['B27'] = str(dpi)
        ws['J24'] = "Asistencia Técnica"

        for hoja in wb.sheetnames:
            if hoja != "ASISTENCIADAUAD1":
                wb.remove(wb[hoja])

        
        ruta_guardado = obtener_ruta_planillas(dep, codigo_oficio)
        nombre_archivo = f"{dep}, ASISTENCIA ADICIONAL DAU.xlsx"
        output_excel = os.path.join(ruta_guardado, nombre_archivo)
        pdf_output = output_excel.replace(".xlsx", ".pdf")
        wb.save(output_excel)

        pythoncom.CoInitialize()
        try:
            excel = client.Dispatch("Excel.Application")
            excel.Application.Visible = False
            wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
            wb_pdf.Sheets(1).Select()
            pdf_output = output_excel.replace(".xlsx", ".pdf")
            wb_pdf.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
            wb_pdf.Close(False)
            excel.Quit()
        finally:
            pythoncom.CoUninitialize()

        with open(pdf_output, "rb") as f:
            st.download_button(
                label="Descargar Planilla de Asistencia Adicional DAU PDF",
                data=f,
                file_name=os.path.basename(pdf_output),
                mime="application/pdf"
            )